import re
import sys
import random

import openpyxl #pip install openpyxl
import numpy #pip install numpy
import pyperclip #pip install pyperclip
import matplotlib.pyplot as plt #pip install matplotlib

import methordanalysis as MAI
import NeuralNetwork #pip install torch

from PyQt5 import QtWidgets #pip install PyQt5==5.15.6
from PyQt5 import QtCore
from PyQt5 import QtGui
from PyQt5.QtWebEngineWidgets import * #pip install PyQtWebEngine==5.15.6
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from Form import Ui_MainWindow
from coderForm import Ui_MainWindow as coder_MainWindow


class MultisiriesStackedBarChart(FigureCanvas):
    def __init__(self, parent=None):
        self.fig , self.ax = plt.subplots(1, dpi=100, figsize=(5, 5),sharey=True, facecolor='white')
        super().__init__(self.fig) 


    def SetParams(self, names, values, legends, colors, fontsize=8):
        values = list(zip(*values))
        values = list([list([y for y in x]) for x in values])
        bottomUse = numpy.array(values[0])
        index = names
        colors = list([x[:-1] for x in colors])
        for altIndex in range(len(values)):
            if altIndex > 0:
                self.ax.bar(index, values[altIndex], color=colors[altIndex], bottom=bottomUse)
                bottomUse = bottomUse + numpy.array(values[altIndex])
            else:
                self.ax.bar(index, values[altIndex], color=colors[altIndex])
        self.fig.legend(legends)
        self.fig.suptitle("Влияние значений на выбор", size=fontsize)


class Diagramme(FigureCanvas):
    def __init__(self, parent=None):     
        self.fig , self.ax = plt.subplots(1, dpi=100, figsize=(5, 5),sharey=True, facecolor='white')
        super().__init__(self.fig) 
        self.name = ""
        names = []
        colors = []
        values = []
        self.ax.bar(names, values, color = colors)
        self.fig.suptitle("",size=8)


    def SetParams(self, names, values, colors, fontsize=8):
        self.ax.bar(names, values, color = colors)
        fontsize = 10
        self.fig.suptitle(self.name, size=fontsize)


class coderWin(QtWidgets.QMainWindow):
    def __init__(self, parentWin):
        super(coderWin, self).__init__()
        self.ui = coder_MainWindow()
        self.ui.setupUi(self)
        self.parentWin = parentWin
        self.ui.pushButton.clicked.connect(self.GetCode)
    

    def GetCode(self):
        self.parentWin.CodeInput.Set(self.ui.plainTextEdit.toPlainText())
        self.hide()
        self.parentWin.RunCodeInText(self.parentWin.CodeInput.Get())


class mywindow(QtWidgets.QMainWindow):
        def __init__(self):
            super(mywindow, self).__init__()
            self.ui = Ui_MainWindow()
            self.ui.setupUi(self)
            self.CodeInput = MAI.CodeInputer()
            self.MethodMAIN = MAI.MAI()
            self.coderwindow = coderWin(self)

            self.ui.tableWidget_3.contextMenuEvent = self.ForMPETableMenuEvent
            self.MethodMAIN.InterfaceVar = {"_UI": self}
            self.ParametersMPE = {}
            self.AlternativesMPE = {}
            self.TempTableOfAlternatives = {}
            self.AnwserMoreInfo = []
            self.AnwserInfo = []
            self.AnwserColors = {}  
            self.ui.Browser.setUrl(QtCore.QUrl("https://www.google.com"))
            self.ui.Browser.urlChanged.connect(lambda x: self.ui.lineEdit.setText(x.toString()))
            self.MainInfluenceGraph = Diagramme()
            self.InALternativeIOnIfluence = MultisiriesStackedBarChart()

            self.ui.widget = QtWidgets.QVBoxLayout(self.ui.widget)
            self.ui.widget.setGeometry(QtCore.QRect(10, 70, 821, 831))
            self.ui.widget.setObjectName("widget")
            self.ui.widget.addWidget(self.MainInfluenceGraph)

            self.NeuralNetworkModel = None
            self.NeuralNetworkTrain = None

            self.DeviceUse = NeuralNetwork.Torch.is_cuda() 
            if self.DeviceUse:
                self.ui.label_23.setText("Cuda присутсвует")
                self.DeviceUse = "cuda:0"
            else:
                self.DeviceUse = "cpu"
                self.ui.label_23.setText("Cuda отсутсвует")
                self.ui.comboBox_6.setEnabled(False)
                self.ui.comboBox_7.setEnabled(False)

            self.SetSwitchButtons()


        def SaveNNModel(self):
            try:
                file = QtWidgets.QFileDialog.getExistingDirectory(None, "Выберите папку куда сохранить MNN")   
                model_body = file+"/model_param.mnn"
                NeuralNetwork.torch.save(self.NeuralNetworkModel.state_dict(), model_body)
                model_init = file+"/model_init.mnn"
                text_to_mnn = f"""{self.NeuralNetworkModel.inout[0]}\n{self.NeuralNetworkModel.inout[1]}\n{self.NeuralNetworkModel.layers_hidden_count}\n{self.NeuralNetworkModel.hidden_layers_input_output}\n{self.NeuralNetworkTrain.mddloss}\n{self.NeuralNetworkTrain.optimizer_name}\n{self.NeuralNetworkTrain.criterion_name}"""
                with open(model_init, "w") as FILE:
                    FILE.write(text_to_mnn)

                self.ui.plainTextEdit_3.setPlainText("Сохранены файлы в: "+str(file))
            except Exception as exp:
                print(exp)


        def LoadNNModel(self):
            try:
                file = QtWidgets.QFileDialog.getExistingDirectory(None, "Выберите папку с сохранённой MNN")   
                with open(file+"/model_init.mnn", "r") as FILE:
                    text_in_mnn = FILE.read()

                arguments = text_in_mnn.split("\n")
                arguments = list([int(arguments[x]) if x < 4 else arguments[x] for x in range(len(arguments))])
                arguments[4] = float(arguments[4])

                self.ui.tableWidget_5.setColumnCount(arguments[0])
                self.ui.tableWidget_6.setColumnCount(arguments[1])

                for i in range(arguments[0]):
                    self.ui.tableWidget_5.setItem(0, i, QtWidgets.QTableWidgetItem(str(i)))
            
                for i in range(arguments[1]):
                    self.ui.tableWidget_6.setItem(0, i, QtWidgets.QTableWidgetItem(str(i)))

                self.ui.label_33.setText(str(round(arguments[4], 7)))

                self.NeuralNetworkModel = NeuralNetwork.NeuralNetwork(arguments[0], arguments[1], arguments[2], arguments[3])
                self.NeuralNetworkModel.load_state_dict(NeuralNetwork.torch.load(file+"/model_param.mnn"))
                self.NeuralNetworkTrain = NeuralNetwork.Training(self.NeuralNetworkModel, arguments[5], 0.001, arguments[6])
                self.NeuralNetworkTrain.midloss = arguments[4]
                self.ui.lineEdit_2.setText(str(arguments[0]))
                self.ui.lineEdit_3.setText(str(arguments[1]))
                self.ui.comboBox_10.setCurrentIndex(arguments[2])
                self.ui.lineEdit_4.setText(str(arguments[3]))

                self.ui.label_27.setText(str(arguments[2]))
                self.ui.label_29.setText(str(arguments[0] + arguments[1] + arguments[2]*arguments[3]))
                self.ui.label_31.setText(str(arguments[3]**(2*(arguments[2] - 1)) + arguments[3]*(arguments[0] + arguments[1])))
                self.ui.label_44.setText("Да")
                self.ui.plainTextEdit_3.setPlainText(str("Загружены файлы из: " + str(file)))
            except Exception as exp:
                print(exp)


        def ExecNNModel(self):
            input_count = self.NeuralNetworkModel.inout[0]
            input_value = []

            for i in range(input_count):
                Item = self.ui.tableWidget_5.item(1, i)
                if Item is None:
                    print(Item, "is Nonetype")
                    return
                input_value.append(float(Item.text()))

            input_value = NeuralNetwork.torch.tensor(input_value)

            output = self.NeuralNetworkModel(input_value)
            indx = (output == output.max().item()).nonzero(as_tuple=True)[0]
            self.ui.label_45.setText("Параметр по индексом: "+str(indx)+", является предсказанным значением нейросети ("+str(round(output.max().item(), 5))+")")

            for i in range(self.NeuralNetworkModel.inout[1]):
                self.ui.tableWidget_6.setItem(1, i, QtWidgets.QTableWidgetItem("0"))
                Item = self.ui.tableWidget_6.item(1, i)
                Item.setText(str(round(output[i].item(), 3)))


        def TrainingBaseLoad(self):
            try:
                file, check = QtWidgets.QFileDialog.getOpenFileName(None, "Открыть neural network base файл", "", "NNB (*.nnb)")
                self.ui.plainTextEdit_4.setPlainText(file)
                with open(file, "r") as FILE:
                    self.NeuralNetworkTrain.base_file = FILE.read()
            except Exception as exp:
                print(exp)
            

        def StartTraining_nn_model(self):
            if self.NeuralNetworkTrain.base_file is None:
                print("NNB файл не загружен")
                return

            epochs = int(self.ui.lineEdit_6.text())
            input_val, target_val = self.NeuralNetworkTrain.get_base()
            self.NeuralNetworkTrain.set_train_val(input_val, target_val)
            losses = self.NeuralNetworkTrain.train(epochs)
            self.ui.listWidget_2.clear()
            
            self.NeuralNetworkTrain.mddloss = sum(losses)/len(losses)
            ldiv = 1 if len(losses) // 50 == 0 else len(losses) // 50
            self.ui.label_33.setText(str(round(self.NeuralNetworkTrain.mddloss, 7)))
            
            com_losses = list([str(str(x)) + ": " +str(round(losses[x], 5)) for x in range(0, len(losses), ldiv)])
            self.ui.listWidget_2.addItems(com_losses)


        def Create_nn_model(self):
            input_ = int(self.ui.lineEdit_2.text())
            output_ = int(self.ui.lineEdit_3.text())
            optim = self.ui.comboBox_5.currentText()
            hidden_layers = int(self.ui.comboBox_10.currentText())
            hidden_layers_input_output = int(self.ui.lineEdit_4.text())

            activation = (self.ui.comboBox_4.currentText())
            criterion = "MSELoss"

            self.ui.label_27.setText(str(hidden_layers))
            self.ui.label_29.setText(str(input_+output_+hidden_layers*hidden_layers_input_output))
            self.ui.label_31.setText(str((hidden_layers_input_output*hidden_layers_input_output)**(hidden_layers-1)+(input_+output_)*hidden_layers_input_output))
            self.ui.label_35.setText(optim)

            self.ui.tableWidget_5.setColumnCount(input_)
            self.ui.tableWidget_6.setColumnCount(output_)

            for i in range(input_):
                self.ui.tableWidget_5.setItem(0, i, QtWidgets.QTableWidgetItem(str(i)))
            
            for i in range(output_):
                self.ui.tableWidget_6.setItem(0, i, QtWidgets.QTableWidgetItem(str(i)))

            self.NeuralNetworkModel = NeuralNetwork.NeuralNetwork(input_, output_, hidden_layers, hidden_layers_input_output)
            self.NeuralNetworkTrain = NeuralNetwork.Training(self.NeuralNetworkModel, optim, 0.0001, criterion)
            self.ui.label_44.setText("Да")


        def RunCodeInText(self, code):
            try: 
                TheMainCmd = code.strip()
                AllVariables = re.findall(r"\|.*?\|", code) #r'\[\[.*?]]\);'
                VariablePosition = list([int(x.split(":")[0].replace("|", ""))-1, int(x.split(":")[1].replace("|", ""))-1, x]  for x in AllVariables)

                for varPos in VariablePosition:
                    try: TheMainCmd = TheMainCmd.replace(varPos[2], self.ui.tableWidget_3.item(varPos[0], varPos[1]).text())
                    except: continue
                self.CodeInput.Ret(self.MethodMAIN.exec_with_return(TheMainCmd)) 
                self.ui.plainTextEdit_2.setPlainText(str(self.CodeInput.Ret()))

            except Exception as exp: 
                try: self.ui.plainTextEdit_2.setPlainText(exp)  
                except: print(exp)         


        def setFontTab(self, table, spin, combobox):
            Cells = table.selectedIndexes()
            fontsize = spin.value()
            fontName = combobox.currentText()
            for cell in Cells:
                row = cell.row()
                column = cell.column()  
                cell = table.item(row, column)
                if cell is None: continue
                font = QtGui.QFont(fontName, fontsize)
                cell.setFont(font)


        def saveNewExcel(self):
            file, check = QtWidgets.QFileDialog.getSaveFileName(None, "Создать excel файл", "", "Excel (*.xlsx)")
            if not(check): return False
            FirstTable = []
            for y in range(self.ui.tableWidget_3.rowCount()):
                FirstTable.append([])
                for x in range(self.ui.tableWidget_3.columnCount()):
                    FirstTable[y].append(self.ui.tableWidget_3.item(y, x).text() if self.ui.tableWidget_3.item(y, x) != None else "")
            
            SecondTable = []
            for y in range(self.ui.tableWidget_2.rowCount()):
                SecondTable.append([])
                for x in range(self.ui.tableWidget_2.columnCount()):
                    SecondTable[y].append(self.ui.tableWidget_2.item(y, x).text() if self.ui.tableWidget_2.item(y, x) != None else "")
            
            SoMuchTables = []
            for param in self.TempTableOfAlternatives:
                SoMuchTables.append([param, self.TempTableOfAlternatives[param]])
            
            AnwserTable = []
            for y in range(self.ui.tableWidget.rowCount()):
                AnwserTable.append([])
                for x in range(2):
                    AnwserTable[y].append(self.ui.tableWidget.item(y, x).text() if self.ui.tableWidget.item(y, x) != None else "")

            book = openpyxl.Workbook(file) 
            TheMainBody = book.create_sheet("The Main Part") 
            for row in FirstTable:
                TheMainBody.append(row)

            TheEndBody = book.create_sheet("The End Part") 
            for row in AnwserTable:
                TheEndBody.append(row)

            TheParametrsBody = book.create_sheet("The Parameters Part") 
            for row in SecondTable:
                TheParametrsBody.append(row)
            
            for alt in SoMuchTables:
                TheTempBody = book.create_sheet("The Parameter - "+str(alt[0])) 
                for row in alt[1]:
                    TheTempBody.append(row)
            book.save(file)


        def ChangeGraphs(self):
            NameOfparametr = self.ui.comboBox_2.currentText()
            plt.close('all')
            if str(NameOfparametr) == "Конечный результат":
                try:
                    self.ui.widget.removeWidget(self.MainInfluenceGraph)
                    self.MainInfluenceGraph.deleteLater()
                    self.MainInfluenceGraph.hide()
                except:
                    pass
                try:
                    self.ui.widget.removeWidget(self.InALternativeIOnIfluence)
                    self.InALternativeIOnIfluence.deleteLater()
                    self.InALternativeIOnIfluence.hide()
                except:
                    pass
                self.MainInfluenceGraph = Diagramme()
                self.MainInfluenceGraph.name = "Конечный результат"

                BigVector = sum(self.AnwserInfo)
                OneProcent = BigVector/100
                values = []
                colors = []
                for value in self.AnwserInfo:
                    values.append(round((value/OneProcent)))
                
                for head in self.HeaderTable[1]:
                    if not(head in self.AnwserColors):
                        tempColor = [random.randint(80, 200)/255, random.randint(80, 200)/255, random.randint(80, 200)/255, 1]
                        self.AnwserColors[head] = tempColor
                        colors.append(tempColor)
                    else:
                        tempColor = self.AnwserColors[head]
                        colors.append(tempColor)

                self.MainInfluenceGraph.SetParams(self.HeaderTable[1], values, colors)
                self.ui.widget.addWidget(self.MainInfluenceGraph)
            elif str(NameOfparametr) == "Параметры":
                try:
                    self.ui.widget.removeWidget(self.MainInfluenceGraph)
                    self.MainInfluenceGraph.deleteLater()
                    self.MainInfluenceGraph.hide()
                except:
                    pass
                try:
                    self.ui.widget.removeWidget(self.InALternativeIOnIfluence)
                    self.InALternativeIOnIfluence.deleteLater()
                    self.InALternativeIOnIfluence.hide()
                except:
                    pass
                
                self.MainInfluenceGraph = Diagramme()
                self.MainInfluenceGraph.name = "Параметры"
                Anwser = self.MethodMAIN.NormalizedVectorsOfMainMatrix
                BigVector = sum(Anwser)
                OneProcent = BigVector/100
                values = []
                colors = []
                for value in Anwser:
                    values.append(round((value/OneProcent)))

                for head in self.HeaderTable[0]:
                    if not(head in self.AnwserColors):
                        tempColor = [random.randint(80, 200)/255, random.randint(80, 200)/255, random.randint(80, 200)/255, 1]
                        self.AnwserColors[head] = tempColor
                        colors.append(tempColor)
                    else:
                        tempColor = self.AnwserColors[head]
                        colors.append(tempColor)

                self.MainInfluenceGraph.SetParams(self.HeaderTable[0], values, colors)
                self.ui.widget.addWidget(self.MainInfluenceGraph)
            elif str(NameOfparametr) == "Влияние параметров":
                try:
                    self.ui.widget.removeWidget(self.MainInfluenceGraph)
                    self.MainInfluenceGraph.deleteLater()
                    self.MainInfluenceGraph.hide()
                except:
                    pass
                try:
                    self.ui.widget.removeWidget(self.InALternativeIOnIfluence)
                    self.InALternativeIOnIfluence.deleteLater()
                    self.InALternativeIOnIfluence.hide()
                except:
                    pass

                self.InALternativeIOnIfluence = MultisiriesStackedBarChart()
                
                colors = []
                for head in self.HeaderTable[0]:
                    if not(head in self.AnwserColors):
                        tempColor = [random.randint(80, 200)/255, random.randint(80, 200)/255, random.randint(80, 200)/255, 1]
                        self.AnwserColors[head] = tempColor
                        colors.append(tempColor)
                    else:
                        tempColor = self.AnwserColors[head]
                        colors.append(tempColor)
            
                response = self.AnwserMoreInfo 
                forFunction = []
                for tg in response:
                    forFunction.append(tg)
                
                lblNumber = list([self.HeaderTable[0][x[1]] for x in response[0]])

                forFunction = list([[y[0] for y in x] for x in response])
                SumParms = list([sum(x) for x in forFunction])
                InProcents = list([[round(100*y/SumParms[x]) for y in forFunction[x]] for x in range(0 ,len(forFunction))])
                self.InALternativeIOnIfluence.SetParams(self.HeaderTable[1], InProcents, lblNumber, colors)
                self.ui.widget.addWidget(self.InALternativeIOnIfluence)


        def GetAnwser(self):
            [response, moreInfo] = self.MethodMAIN.GetAnwser(len(self.HeaderTable[1]))
            self.AnwserMoreInfo = moreInfo
            self.AnwserInfo = response
            Headers = self.HeaderTable[1]
            self.ui.tableWidget.setRowCount(len(Headers))
            for y in range(len(Headers)):
                self.ui.tableWidget.setItem(y, 0, QtWidgets.QTableWidgetItem(str(Headers[y])))
                self.ui.tableWidget.setItem(y, 1, QtWidgets.QTableWidgetItem(str(response[y])))
                self.ui.tableWidget.item(y, 1).setBackground(QtGui.QColor(0,0,0,0))
                self.ui.tableWidget.item(y, 0).setBackground(QtGui.QColor(0,0,0,0))
            
            maxIs = max(response)
            maxInd = response.index(maxIs)
            self.ui.tableWidget.item(maxInd, 0).setBackground(QtGui.QColor(255,0,0,100))
            self.ui.tableWidget.item(maxInd, 1).setBackground(QtGui.QColor(255,0,0,100))


        def SaveTableAlternative(self):
            Table = []
            NowAlternative = self.ui.comboBox.currentText()
            for y in range(1, self.ui.tableWidget_4.rowCount()):
                Table.append([])
                for x in range(1, self.ui.tableWidget_4.columnCount()):
                    Table[y-1].append(float(self.ui.tableWidget_4.item(y, x).text()))
            self.TempTableOfAlternatives[NowAlternative] = Table
            [NormalaizedVectors, ConsistencyRelation] = self.MethodMAIN.AnalysisTable(Table)
            self.MethodMAIN.NormalaizedVectorsForParameters[NowAlternative] = [NormalaizedVectors, self.ui.comboBox.currentIndex()]
            string = "больше 10%, -> таблицу использовать НЕЛЬЗЯ" if ConsistencyRelation >= 10 else "меньше 10%, -> таблицу использовать МОЖНО"
            self.ui.label_8.setText(f"Вывод: Отношение соответсвия({ConsistencyRelation})({NowAlternative}) = {string}")


        def AlternativeChanged(self):
            self.clearTableWidget_4()
            self.ui.tableWidget_4.clear()
            NowAlternative = self.ui.comboBox.currentText()
            TableUsed = self.TempTableOfAlternatives[NowAlternative]
            self.ui.tableWidget_4.setRowCount(len(TableUsed)+1)
            self.ui.tableWidget_4.setColumnCount(len(TableUsed)+1)
            self.ui.tableWidget_4.setVerticalHeaderLabels(list([str(x) for x in range(1, len(TableUsed)+2)]))
            self.ui.tableWidget_4.setHorizontalHeaderLabels(list([str(x) for x in range(1, len(TableUsed)+2)]))
            for x in range(1, len(TableUsed)+1):
                self.ui.tableWidget_4.setItem(0, x, QtWidgets.QTableWidgetItem(str(self.HeaderTable[1][x-1])))
                self.ui.tableWidget_4.item(0, x).setBackground(QtGui.QColor(0,0,255,100))
                self.ui.tableWidget_4.setItem(x, 0, QtWidgets.QTableWidgetItem(str(self.HeaderTable[1][x-1])))
                self.ui.tableWidget_4.item(x, 0).setBackground(QtGui.QColor(0,0,255,100))
            for x in range(1, len(TableUsed)+1):
                for y in range(1, len(TableUsed)+1):
                    self.ui.tableWidget_4.setItem(x, y, QtWidgets.QTableWidgetItem(str(TableUsed[x-1][y-1])))
                self.ui.tableWidget_4.setItem(x, x, QtWidgets.QTableWidgetItem("1"))


        def MethodParentEquals(self):
            try:
                Table = []
                for y in range(1, self.ui.tableWidget_2.rowCount()):
                    Table.append([])
                    for x in range(1, self.ui.tableWidget_2.columnCount()):
                        Table[y-1].append(float(self.ui.tableWidget_2.item(y, x).text()))
                [NormalaizedVectors, ConsistencyRelation] = self.MethodMAIN.AnalysisTable(Table)
                self.MethodMAIN.NormalizedVectorsOfMainMatrix = NormalaizedVectors
                string = "больше 10%, -> таблицу использовать НЕЛЬЗЯ" if ConsistencyRelation >= 10 else "меньше 10%, -> таблицу использовать МОЖНО"
                self.ui.label_7.setText(f"Вывод: Отношение соответсвия({ConsistencyRelation}) = {string}")
            except Exception as exp:
                print(exp)         


        def SaveMPETable(self):
            try:
                self.AnwserColors = {}
                self.MethodMAIN.NormalaizedVectorsForParameters = {}
                FiltredTable = []
                Header = [[], []]
                for y in range(len(self.ParametersMPE)):
                    
                    FiltredTable.append([])
                    for _ in range(len(self.AlternativesMPE)):
                        FiltredTable[y].append(None)
                for y in self.ParametersMPE:
                    Header[0].append(y)
                self.MethodMAIN.NameOfAlternative = []
                for x in self.AlternativesMPE:
                    self.MethodMAIN.NameOfAlternative.append(x)
                    Header[1].append(x)
                now_y = 0
                for par in self.ParametersMPE:
                    now_x = 0
                    for alt in self.AlternativesMPE:
                        text = self.ui.tableWidget_3.item(self.ParametersMPE[par][0], self.AlternativesMPE[alt][1])
                        if text == "" or text == None:
                            text == "-"
                        else:
                            text = text.text()
                        FiltredTable[now_y][now_x] = text
                        now_x+=1
                    now_y+=1

                Column = len(Header[0]) 
                Rows = len(Header[1])
                self.ui.comboBox.clear()  
                self.ui.comboBox.addItems(Header[0])
                for alt in Header[0]:
                    self.TempTableOfAlternatives[alt] = []
                    for ind, _ in enumerate(Header[1]):
                        self.TempTableOfAlternatives[alt].append([])
                        for _ in Header[1]:
                            self.TempTableOfAlternatives[alt][ind].append("")
                self.ui.tableWidget_2.setColumnCount(Column+1)
                self.ui.tableWidget_2.setRowCount(Column+1)
                self.ui.tableWidget_2.setVerticalHeaderLabels(list([str(x) for x in range(1, Column+2)]))
                self.ui.tableWidget_2.setHorizontalHeaderLabels(list([str(x) for x in range(1, Column+2)]))

                self.ui.label_3.setText(f"{len(Header[0])}x{len(Header[0])}")
                self.ui.label_5.setText(f"{len(Header[1])}x{len(Header[1])}")

                for x in range(Column):
                    self.ui.tableWidget_2.setItem(0, x+1, QtWidgets.QTableWidgetItem(f"{Header[0][x]}"))
                    self.ui.tableWidget_2.item(0, x+1).setBackground(QtGui.QColor(0,255,0,100))
                    self.ui.tableWidget_2.setItem(x+1, 0, QtWidgets.QTableWidgetItem(f"{Header[0][x]}"))
                    self.ui.tableWidget_2.item(x+1, 0).setBackground(QtGui.QColor(0,255,0,100))
                    self.ui.tableWidget_2.setItem(x+1, x+1, QtWidgets.QTableWidgetItem("1"))

                self.FiltredTable = FiltredTable
                self.HeaderTable = Header

            except Exception as exp:
                print(exp)


        def ForMPETableMenuEvent(self, event):
            try:
                nowcurrentRow, nowcurrentColumn = self.ui.tableWidget_3.currentRow(), self.ui.tableWidget_3.currentColumn()
                menu = QtWidgets.QMenu()
                
                anwser = None
                for alt in self.AlternativesMPE:
                    if self.AlternativesMPE[alt] == [nowcurrentRow, nowcurrentColumn]:
                        anwser = "Alt"
                for par in self.ParametersMPE:      
                    if self.ParametersMPE[par] == [nowcurrentRow, nowcurrentColumn]:
                        anwser = "Par"

                Cells = self.ui.tableWidget_3.selectedIndexes()
                ExecuteTheCommand = menu.addAction("Выполнить")
                CopyIdCell = menu.addAction("Скопировать ID")
                
                if anwser == None:
                    setLikeParameter = menu.addAction("Назначить параметром")
                    setLikeAlternative = menu.addAction("Назначить альтернативой")
                elif anwser == "Alt":
                    delLikeAlternative = menu.addAction("Удалить альтернативу")
                elif anwser == "Par":
                    delLikeParameter = menu.addAction("Удалить Параметр")

                action = menu.exec_(self.mapToGlobal(event.pos()))

                if action == ExecuteTheCommand:
                    command = self.ui.tableWidget_3.item(nowcurrentRow, nowcurrentColumn).text()
                    self.execute_cmd(command, [nowcurrentRow, nowcurrentColumn])
                    return
                elif action == CopyIdCell:
                    string_to_copy = ""
                    for cell in Cells:
                        row = cell.row()
                        column = cell.column()  
                        string_to_copy += f"|{row+1}:{column+1}|,"
                    pyperclip.copy(string_to_copy)
                    return

                for cell in Cells:
                    row = cell.row()
                    column = cell.column()              
                    if anwser == None:
                        if action == setLikeParameter:
                            self.ParametersMPE[self.ui.tableWidget_3.item(row, column).text()] = [row, column]
                            self.ui.tableWidget_3.item(row, column).setBackground(QtGui.QColor(0,255,0,100))
                            ItemCell = self.ui.tableWidget_3.item(row, column)    
                            ItemCell.setFlags(ItemCell.flags() & ~QtCore.Qt.ItemIsEditable)
                        elif action == setLikeAlternative:
                            self.AlternativesMPE[self.ui.tableWidget_3.item(row, column).text()] = [row, column]
                            self.ui.tableWidget_3.item(row, column).setBackground(QtGui.QColor(0,0,255,100))
                            ItemCell = self.ui.tableWidget_3.item(row, column)    
                            ItemCell.setFlags(ItemCell.flags() & ~QtCore.Qt.ItemIsEditable)
                    else:
                        if anwser == "Alt":
                            if action == delLikeAlternative:
                                self.ui.tableWidget_3.item(row, column).setBackground(QtGui.QColor(0,0,0,0))
                                del self.AlternativesMPE[self.ui.tableWidget_3.item(row, column).text()]
                                ItemCell = self.ui.tableWidget_3.item(row, column)
                                ItemCell.setFlags(ItemCell.flags() | QtCore.Qt.ItemIsEditable)
                        elif anwser == "Par":
                            if action == delLikeParameter:
                                self.ui.tableWidget_3.item(row, column).setBackground(QtGui.QColor(0,0,0,0))
                                del self.ParametersMPE[self.ui.tableWidget_3.item(row, column).text()]
                                ItemCell = self.ui.tableWidget_3.item(row, column)
                                ItemCell.setFlags(ItemCell.flags() | QtCore.Qt.ItemIsEditable)
            except Exception as exp:
                if "text" in str(exp):
                    print("Не введено имя")
                else:
                    print(exp)


        def execute_cmd(self, cmd, positions):
            try:
                TheMainCmd = cmd.strip()
                AllVariables = re.findall(r"\|.*?\|", cmd) #r'\[\[.*?]]\);'
                VariablePosition = list([int(x.split(":")[0].replace("|", ""))-1, int(x.split(":")[1].replace("|", ""))-1, x]  for x in AllVariables)
                for varPos in VariablePosition:
                    TheMainCmd = TheMainCmd.replace(varPos[2], self.ui.tableWidget_3.item(varPos[0], varPos[1]).text())
                try: anwser = round(eval(TheMainCmd), 4)
                except: anwser = TheMainCmd
                self.ui.tableWidget_3.setItem(positions[0], positions[1], QtWidgets.QTableWidgetItem(str(anwser)))
            except Exception as exp:
                print(exp)     


        def SetSwitchButtons(self):
            self.ui.pushButton_54.clicked.connect(self.LoadNNModel)
            self.ui.pushButton_48.clicked.connect(self.SaveNNModel)
            self.ui.pushButton_55.clicked.connect(self.ExecNNModel)
            self.ui.pushButton_49.clicked.connect(self.TrainingBaseLoad)
            self.ui.pushButton_53.clicked.connect(self.StartTraining_nn_model)
            self.ui.pushButton_46.clicked.connect(self.Create_nn_model)

            self.ui.pushButton_27.clicked.connect(self.coderwindow.show)
            self.ui.pushButton_10.clicked.connect(self.SaveMPETable)
            self.ui.pushButton_43.clicked.connect(self.saveNewExcel)
            self.ui.pushButton_11.clicked.connect(lambda : self.setFontTab(self.ui.tableWidget_3, self.ui.spinBox_3, self.ui.fontComboBox))
            self.ui.pushButton_12.clicked.connect(lambda : self.setFontTab(self.ui.tableWidget_2, self.ui.spinBox_4, self.ui.fontComboBox_2))
            self.ui.pushButton_13.clicked.connect(lambda : self.setFontTab(self.ui.tableWidget_4, self.ui.spinBox_5, self.ui.fontComboBox_3))

            self.ui.tableWidget_2.cellChanged.connect(self.ChangeNumbers_1)
            self.ui.tableWidget_4.cellChanged.connect(self.ChangeNumbers_2)
            self.ui.pushButton_29.clicked.connect(self.SaveTableAlternative)
            self.ui.pushButton_44.clicked.connect(self.GetAnwser)
            self.ui.comboBox.activated.connect(self.AlternativeChanged)
            self.ui.comboBox_2.activated.connect(self.ChangeGraphs)

            self.ui.spinBox.valueChanged.connect(lambda x: [self.ui.tableWidget_3.setRowCount(x) ,self.ui.tableWidget_3.setVerticalHeaderItem(x-1, QtWidgets.QTableWidgetItem(f"{x}"))])
            self.ui.spinBox_2.valueChanged.connect(lambda x: [self.ui.tableWidget_3.setColumnCount(x) ,self.ui.tableWidget_3.setHorizontalHeaderItem(x-1, QtWidgets.QTableWidgetItem(f"{x}"))])

            self.ui.pushButton.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(0))
            self.ui.pushButton_4.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(1))
            self.ui.pushButton_7.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(2))
            self.ui.pushButton_26.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(3))
            self.ui.pushButton_39.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(4))
            self.ui.pushButton_45.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(5))

            self.ui.pushButton_2.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(0))
            self.ui.pushButton_3.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(1))
            self.ui.pushButton_8.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(2))
            self.ui.pushButton_25.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(3))
            self.ui.pushButton_38.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(4))
            self.ui.pushButton_61.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(5))

            self.ui.pushButton_5.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(0))
            self.ui.pushButton_6.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(1))
            self.ui.pushButton_9.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(2))
            self.ui.pushButton_24.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(3))
            self.ui.pushButton_37.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(4))
            self.ui.pushButton_47.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(5))

            self.ui.pushButton_20.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(0))
            self.ui.pushButton_21.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(1))
            self.ui.pushButton_22.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(2))
            self.ui.pushButton_23.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(3))
            self.ui.pushButton_36.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(4))
            self.ui.pushButton_51.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(5))

            self.ui.pushButton_31.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(0))
            self.ui.pushButton_32.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(1))
            self.ui.pushButton_33.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(2))
            self.ui.pushButton_34.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(3))
            self.ui.pushButton_35.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(4))
            self.ui.pushButton_50.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(5))

            self.ui.pushButton_17.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(0))
            self.ui.pushButton_18.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(1))
            self.ui.pushButton_19.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(2))
            self.ui.pushButton_28.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(3))
            self.ui.pushButton_41.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(4))
        
            self.ui.pushButton_40.clicked.connect(lambda: self.ui.Browser.setUrl(QtCore.QUrl("https://www.google.com")))
            self.ui.pushButton_16.clicked.connect(self.MethodParentEquals)
            self.ui.pushButton_15.clicked.connect(self.clearTableWidget_2)
            self.ui.pushButton_14.clicked.connect(self.clearTableWidget_3)
            self.ui.pushButton_30.clicked.connect(self.clearTableWidget_4)

            self.ui.pushButton_57.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(1))
            self.ui.pushButton_58.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(1))
            self.ui.pushButton_59.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(0))
            self.ui.pushButton_60.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(0))


        def clearTableWidget_2(self):
            for x in range(1, self.ui.tableWidget_2.rowCount()):
                for y in range(1, self.ui.tableWidget_2.columnCount()):
                    self.ui.tableWidget_2.setItem(x, y, QtWidgets.QTableWidgetItem(""))


        def clearTableWidget_3(self):
            for x in range(1, self.ui.tableWidget_2.rowCount()):
                for y in range(1, self.ui.tableWidget_2.columnCount()):
                    self.ui.tableWidget_2.setItem(x, y, QtWidgets.QTableWidgetItem(""))


        def clearTableWidget_4(self):
            for x in range(1, self.ui.tableWidget_4.rowCount()):
                for y in range(1, self.ui.tableWidget_4.columnCount()):
                    self.ui.tableWidget_4.setItem(x, y, QtWidgets.QTableWidgetItem(""))


        def ChangeNumbers_1(self, col, row):
                    try:
                        cell = self.ui.tableWidget_2.item(col, row)
                        if not(cell == None or row == col):
                            text = cell.text()
                            if (text in "123456789" and not(text == "")):
                                anwser = round(1/int(text), 7)
                                self.ui.tableWidget_2.setItem(row, col, QtWidgets.QTableWidgetItem(str(anwser)))
                    except Exception as exp:
                        print(exp)       


        def ChangeNumbers_2(self, col, row):
            try:
                cell = self.ui.tableWidget_4.item(col, row)
                if not(cell == None or row == col):
                    text = cell.text()
                    if (text in "123456789" and not(text == "")):
                        anwser = round(1/int(text), 7)
                        self.ui.tableWidget_4.setItem(row, col, QtWidgets.QTableWidgetItem(str(anwser)))
            except Exception as exp:
                print(exp)       


def main(): 
    app = QtWidgets.QApplication([])
    application = mywindow()
    
    application.setFixedSize(1250, 905)
    application.show()
    app.setWindowIcon(QtGui.QIcon('./icons/icon.ico'))
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
